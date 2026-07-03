
from pathlib import Path
import json

from openpyxl import load_workbook
from rapidfuzz import process

LABELS = Path("models/birds/classifier/convnext_v1_tiny_eu_common_labels.txt")
BIRDBASE = Path("models/birds/classifier/birdbase_v2025.1.xlsx")
OUTPUT = Path("models/birds/classifier/birdbase_mapping.json")
UNMATCHED = Path("models/birds/classifier/birdbase_unmatched.txt")


def normalize(name: str) -> str:

    name = name.lower()

    # Interpunctie
    name = name.replace("-", " ")
    name = name.replace("'", "")

    # Spelling
    name = name.replace("grey", "gray")

    # Woorden die vaak verschillen tussen taxonomieën
    prefixes = [
        "asian ",
        "atlantic ",
        "western ",
        "eastern ",
        "northern ",
        "southern ",
        "greater ",
        "lesser ",
    ]

    for prefix in prefixes:
        if name.startswith(prefix):
            name = name[len(prefix):]

    return " ".join(name.split())


print("BirdBase laden...")

wb = load_workbook(BIRDBASE, data_only=True)
ws = wb["Data"]

birdbase = {}
normalized = {}

for row in ws.iter_rows(min_row=3, values_only=True):

    english = str(row[1]).strip()

    birdbase[english] = english
    normalized[normalize(english)] = english

print(f"{len(birdbase)} BirdBase soorten geladen.\n")

mapping = {}
missing = []

labels = LABELS.read_text(encoding="utf-8").splitlines()

for label in labels:

    label = label.strip()

    if not label or label == "Unknown":
        continue

    # ----------------------------------------------------------
    # Exacte match
    # ----------------------------------------------------------

    if label in birdbase:

        mapping[label] = label
        print(f"✅ {label}")

        continue

    # ----------------------------------------------------------
    # Genormaliseerde match
    # ----------------------------------------------------------

    norm = normalize(label)

    if norm in normalized:

        bird = normalized[norm]

        mapping[label] = bird

        print(f"🟢 {label}")
        print(f"   -> {bird}")

        continue

    # ----------------------------------------------------------
    # Fuzzy (alleen zeer betrouwbare matches)
    # ----------------------------------------------------------

    match = process.extractOne(
        norm,
        normalized.keys(),
        score_cutoff=99
    )

    if match:

        bird = normalized[match[0]]

        mapping[label] = bird

        print(f"🟡 {label}")
        print(f"   -> {bird} ({match[1]:.1f}%)")

    else:

        print(f"⚪ {label}")

        missing.append(label)

print()
print("=" * 60)
print(f"Labels        : {len(labels)}")
print(f"Gematcht      : {len(mapping)}")
print(f"Niet gevonden : {len(missing)}")
print("=" * 60)

print(f"\nMapping opgeslagen : {OUTPUT}")
print(f"Niet gevonden      : {UNMATCHED}")

OUTPUT.write_text(
    json.dumps(
        dict(sorted(mapping.items())),
        indent=4,
        ensure_ascii=False
    ),
    encoding="utf-8"
)