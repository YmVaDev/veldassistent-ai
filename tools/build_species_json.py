
```python
from pathlib import Path
import json

from openpyxl import load_workbook

LABELS = Path("models/birds/classifier/labels.txt")
BIRDBASE = Path("models/birds/classifier/birdbase_v2025.1.xlsx")
MAPPING = Path("models/birds/classifier/birdbase_mapping.json")
OUTPUT = Path("models/birds/classifier/species.json")

print("Load BirdBase...")

wb = load_workbook(BIRDBASE, data_only=True)
ws = wb["Data"]

# ---------------------------------------------------------------------
# BirdBase in geheugen laden
# ---------------------------------------------------------------------

birdbase = {}

for row in ws.iter_rows(min_row=3, values_only=True):

    english = str(row[1]).strip()

    birdbase[english] = {
        "birdbaseId": row[0],
        "scientificName": row[2],
        "habitat": row[48],
        "diet": row[50],
        "priority": "normal",
        "clip_duration": 0,
        "count": 0,
        "best_score": 0,
        "first_seen": None,
        "last_seen": None
    }

print(f"{len(birdbase)} BirdBase species loaded")

# ---------------------------------------------------------------------
# Mapping laden
# ---------------------------------------------------------------------

mapping = json.loads(
    MAPPING.read_text(encoding="utf-8")
)

# ---------------------------------------------------------------------
# species.json bouwen
# ---------------------------------------------------------------------

species = {}

matched = 0
missing = []

labels = LABELS.read_text(encoding="utf-8").splitlines()

for i, label in enumerate(labels, start=1):

    label = label.strip()

    if not label:
        continue

    print(f"[{i}/{len(labels)}] {label}")

    if label == "Unknown":

        species[label] = {
            "birdbaseId": None,
            "scientificName": None,
            "habitat": None,
            "diet": None,
            "priority": "normal",
            "clip_duration": 0,
            "count": None,
            "best_score": None,
            "first_seen": None,
            "last_seen": None
        }

        continue

    # Gebruik de mapping indien aanwezig
    birdbase_name = mapping.get(label, label)

    data = birdbase.get(birdbase_name)

    if data:

        species[label] = data
        matched += 1

        print(f"   ✅ {data['scientificName']}")

    else:

        print("   ❌ Niet gevonden")

        species[label] = {
            "birdbaseId": None,
            "scientificName": None,
            "habitat": None,
            "diet": None,
            "priority": "normal",
            "clip_duration": 0,
            "count": None,
            "best_score": None,
            "first_seen": None,
            "last_seen": None
        }

        missing.append(label)

# ---------------------------------------------------------------------
# Opslaan
# ---------------------------------------------------------------------

OUTPUT.parent.mkdir(parents=True, exist_ok=True)

OUTPUT.write_text(
    json.dumps(
        species,
        indent=4,
        ensure_ascii=False
    ),
    encoding="utf-8"
)

print()
print("================================")
print(f"Species loaded : {len(species)}")
print(f"Found          : {matched}")
print(f"Not found      : {len(missing)}")
print(f"Saved to JSON  : {OUTPUT}")
print("================================")

if missing:

    print("\nNot found:")

    for s in missing:
        print(" -", s)
```
